import glob
import multiprocessing
import os
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime

import jinja2
import pandas as pd
from docxtpl import DocxTemplate
import openpyxl

from utils import format_date, floatformat, is_date_string


def smart_read_excel(file_path, log_callback=None):
    """
    Smart Excel reader that:
    - Preserves leading zeros for numeric codes
    - Properly handles dates
    - Automatically determines data types
    """
    if log_callback:
        log_callback(f"🔍 Analyzing file structure: {os.path.basename(file_path)}")

    # Step 1: Analyze source file using openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active

    # Get headers
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())
        else:
            headers.append(f"Column_{len(headers)}")

    # Analyze each column
    text_columns = []
    date_columns = []

    for col_idx, header in enumerate(headers, 1):
        sample_values = []

        # Take first 20 non-empty values for analysis
        for row_idx in range(2, min(ws.max_row + 1, 22)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                sample_values.append(cell.value)
            if len(sample_values) >= 10:  # 10 values is enough for analysis
                break

        if not sample_values:
            continue

        # Check for leading zeros
        has_leading_zeros = False
        for val in sample_values[:5]:  # Check first 5 values
            if isinstance(val, str):
                # If string starts with 0 and contains only digits
                if (val.startswith('0') and len(val) > 1 and val.isdigit()):
                    has_leading_zeros = True
                    break
            elif isinstance(val, (int, float)):
                # Check if this was originally a string with leading zero
                # This can be determined by cell format
                cell = ws.cell(row=2, column=col_idx)  # First data row
                if cell.number_format.startswith('0') or cell.number_format == '@':
                    has_leading_zeros = True
                    break

        # Check for dates
        is_date_column = False
        if not has_leading_zeros:  # If this is not a leading zeros column
            for val in sample_values[:3]:
                if isinstance(val, datetime):
                    is_date_column = True
                    break
                elif isinstance(val, str):
                    # Check strings for date similarity
                    if any(word in header.lower() for word in ['date', 'дата', 'время', 'time', 'created', 'updated']):
                        is_date_column = True
                        break
                    # Or by content
                    if is_date_string(val):
                        is_date_column = True
                        break

        # Save analysis results
        if has_leading_zeros:
            text_columns.append(header)
        elif is_date_column:
            date_columns.append(header)

    if log_callback:
        if text_columns:
            log_callback(f"🔢 Columns with leading zeros: {text_columns}")
        if date_columns:
            log_callback(f"📅 Date columns: {date_columns}")

    # Step 2: Read file with correct types
    dtype_dict = {}

    # Read leading zeros columns as strings
    for col in text_columns:
        dtype_dict[col] = str

    # Prepare parameters for pandas
    read_params = {
        'io': file_path,
        'dtype': dtype_dict if dtype_dict else None
    }

    # If there are date columns, specify them for parsing
    if date_columns:
        read_params['parse_dates'] = date_columns

    # Read file
    df = pd.read_excel(**read_params)
    df.columns = df.columns.str.strip()

    return df


def get_optimal_workers():
    """Returns optimal number of worker processes (half of CPU cores)"""
    cpu_count = multiprocessing.cpu_count()
    # Use half of cores, but minimum 2 and maximum 8 for stability
    optimal_workers = max(2, min(8, cpu_count // 2))
    return optimal_workers


def process_single_document(args):
    """
    Function for processing single document in separate process.
    Must be at module top level for pickle serialization.
    """
    try:
        (row_data, template_path, output_dir, common_column,
         file_name_column, other_tables, main_columns) = args

        index, borrower_dict = row_data

        # Create Jinja2 environment in each process
        jinja_env = jinja2.Environment()

        # Define filters directly here to avoid import issues
        def dateonly_filter(val):
            """Date only without time"""
            if pd.isnull(val):
                return '—'

            try:
                # First try to convert everything to pandas datetime
                parsed_date = None

                # If it's already a datetime object
                if hasattr(val, 'strftime'):
                    parsed_date = val
                # If it's a string
                elif isinstance(val, str):
                    if is_date_string(val):
                        parsed_date = pd.to_datetime(val, errors='coerce')
                    else:
                        return str(val)  # Doesn't look like a date
                # If it's a number (timestamp)
                elif isinstance(val, (int, float)):
                    parsed_date = pd.to_datetime(val, unit='s', errors='coerce')
                else:
                    # Try to convert anything else
                    parsed_date = pd.to_datetime(val, errors='coerce')

                # If successfully parsed date
                if parsed_date is not None and pd.notna(parsed_date):
                    # Format ONLY date without time
                    if hasattr(parsed_date, 'date'):
                        return parsed_date.date().strftime('%d.%m.%Y')
                    else:
                        return parsed_date.strftime('%d.%m.%Y')

                return str(val)

            except Exception as e:
                return str(val)

        def datetime_full_filter(val):
            """Date with time"""
            if pd.isnull(val):
                return '—'

            try:
                # Pandas Timestamp or any object with strftime method
                if hasattr(val, 'strftime'):
                    # Check if there's time
                    if hasattr(val, 'time') and val.time() != datetime.min.time():
                        return val.strftime('%d.%m.%Y %H:%M:%S')
                    else:
                        return val.strftime('%d.%m.%Y')

                # Date string
                if isinstance(val, str):
                    if is_date_string(val):
                        parsed_date = pd.to_datetime(val, errors='coerce')
                        if pd.notna(parsed_date):
                            if parsed_date.time() != datetime.min.time():
                                return parsed_date.strftime('%d.%m.%Y %H:%M:%S')
                            else:
                                return parsed_date.strftime('%d.%m.%Y')
                    return str(val)

                return str(val)

            except Exception as e:
                return str(val)

        def datetime_full_no_sec_filter(val):
            """Date with time without seconds"""
            if pd.isnull(val):
                return '—'

            try:
                # Pandas Timestamp or any object with strftime method
                if hasattr(val, 'strftime'):
                    # Check if there's time
                    if hasattr(val, 'time') and val.time() != datetime.min.time():
                        return val.strftime('%d.%m.%Y %H:%M')
                    else:
                        return val.strftime('%d.%m.%Y')

                # Date string
                if isinstance(val, str):
                    if is_date_string(val):
                        parsed_date = pd.to_datetime(val, errors='coerce')
                        if pd.notna(parsed_date):
                            if parsed_date.time() != datetime.min.time():
                                return parsed_date.strftime('%d.%m.%Y %H:%M')
                            else:
                                return parsed_date.strftime('%d.%m.%Y')
                    return str(val)

                return str(val)

            except Exception as e:
                return str(val)

        def number_thousands_filter(val):
            """Number with thousands separators"""
            try:
                if isinstance(val, str):
                    val = val.replace(' ', '').replace(',', '.')
                num = float(val)
                formatted = f"{num:.2f}".replace('.', ',')
                parts = formatted.split(',')
                integer_part = parts[0]
                decimal_part = parts[1] if len(parts) > 1 else ''
                integer_with_spaces = ''
                for i, digit in enumerate(reversed(integer_part)):
                    if i > 0 and i % 3 == 0:
                        integer_with_spaces = ' ' + integer_with_spaces
                    integer_with_spaces = digit + integer_with_spaces
                if decimal_part:
                    return integer_with_spaces + ',' + decimal_part
                else:
                    return integer_with_spaces
            except Exception:
                return val

        def currency_uah_filter(val):
            """Ukrainian hryvnia currency"""
            try:
                formatted = number_thousands_filter(val)
                return f"{formatted} ₴"
            except Exception:
                return val

        def currency_usd_filter(val):
            """US dollar currency"""
            try:
                formatted = number_thousands_filter(val)
                return f"{formatted} $"
            except Exception:
                return val

        # Register filters
        jinja_env.filters['floatformat'] = floatformat
        jinja_env.filters['dateonly'] = dateonly_filter
        jinja_env.filters['datetime_full'] = datetime_full_filter
        jinja_env.filters['datetime_full_no_sec'] = datetime_full_no_sec_filter
        jinja_env.filters['number_thousands'] = number_thousands_filter
        jinja_env.filters['currency_uah'] = currency_uah_filter
        jinja_env.filters['currency_usd'] = currency_usd_filter

        # Additional date filters (synonyms)
        jinja_env.filters['date'] = dateonly_filter
        jinja_env.filters['dateformat'] = dateonly_filter

        # Prepare context for template
        context = {}

        # Add data from main table
        for col in main_columns:
            val = borrower_dict.get(col)
            key = f"{col}_credit"

            # Special handling for dates
            if val is not None and not pd.isnull(val):
                # If it's a datetime object, keep it as is
                if isinstance(val, (pd.Timestamp, datetime)):
                    context[key] = val
                # If it's a date string
                elif isinstance(val, str) and val != "NaT" and is_date_string(val):
                    try:
                        parsed_date = pd.to_datetime(val, errors='coerce')
                        if pd.notna(parsed_date):
                            context[key] = parsed_date
                        else:
                            context[key] = val
                    except:
                        context[key] = val
                else:
                    context[key] = val
            else:
                context[key] = "—"

        # Add data from additional tables
        for tablename, df_dict in other_tables.items():
            # Restore DataFrame from dictionary
            df = pd.DataFrame(df_dict['data'])
            df.columns = df_dict['columns']

            if common_column not in df.columns:
                continue

            # Filter by common column
            borrower_id = borrower_dict.get(common_column)
            filtered = df[df[common_column] == borrower_id]

            rows = []
            for _, row in filtered.iterrows():
                row_dict = {}
                for col in df.columns:
                    val = row[col]

                    # Special handling for dates
                    if val is not None and not pd.isnull(val):
                        # If it's a datetime object, keep it as is
                        if isinstance(val, (pd.Timestamp, datetime)):
                            row_dict[col] = val
                        # If it's a date string
                        elif isinstance(val, str) and val != "NaT" and is_date_string(val):
                            try:
                                parsed_date = pd.to_datetime(val, errors='coerce')
                                if pd.notna(parsed_date):
                                    row_dict[col] = parsed_date
                                else:
                                    row_dict[col] = val
                            except:
                                row_dict[col] = val
                        else:
                            row_dict[col] = val
                    else:
                        row_dict[col] = "—"
                rows.append(row_dict)
            context[f"{tablename}_table"] = rows

        # Document generation
        tpl = DocxTemplate(template_path)
        tpl.render(context, jinja_env)

        # Create filename
        safe_name = str(borrower_dict.get(file_name_column, borrower_dict.get(common_column, f"doc_{index}"))).replace(
            " ", "_")
        # Remove unsafe characters from filename
        safe_name = "".join(c for c in safe_name if c.isalnum() or c in ('-', '_', '.'))

        docx_filename = os.path.join(output_dir, f"doc_{safe_name}.docx")
        tpl.save(docx_filename)

        return {"success": True, "filename": docx_filename, "index": index}

    except Exception as e:
        return {"success": False, "error": str(e), "index": index}


def generate_documents(root_dir, main_path, template_path, output_dir,
                       common_column, file_name_column, log_callback, stop_flag):
    try:
        # Determine number of worker processes
        max_workers = get_optimal_workers()
        log_callback(f"=== Starting DOCX generation ===")
        log_callback(f"💻 Using {max_workers} processes out of {multiprocessing.cpu_count()} available cores")

        if not all([os.path.exists(main_path), os.path.exists(template_path), os.path.isdir(root_dir)]):
            log_callback("❌ Error: Check all file paths!")
            return

        os.makedirs(output_dir, exist_ok=True)

        # Read main table with smart analysis
        log_callback("📖 Reading main table...")
        main_df = smart_read_excel(main_path, log_callback)

        # Read additional tables
        log_callback("📖 Reading additional tables...")
        all_xlsx = glob.glob(os.path.join(root_dir, "*.xlsx"))
        other_xlsx = [f for f in all_xlsx if os.path.abspath(f) != os.path.abspath(main_path)]
        other_tables = {}

        for fname in other_xlsx:
            if stop_flag():
                log_callback("⛔ Generation stopped by user.")
                return

            name = os.path.splitext(os.path.basename(fname))[0].lower()
            # Smart reading for additional tables
            df = smart_read_excel(fname, log_callback)

            # Convert DataFrame to serializable format
            other_tables[name] = {
                'data': df.to_dict('records'),
                'columns': df.columns.tolist()
            }
            log_callback(f"✓ Loaded table: {name} ({len(df)} records)")

        log_callback(f"📊 Found {len(main_df)} records to process")
        log_callback(f"🚀 Starting {max_workers} parallel processes...")

        # Prepare data for parallel processing
        main_columns = main_df.columns.tolist()

        # Convert rows to dictionaries for serialization
        tasks = []
        for i, (_, row) in enumerate(main_df.iterrows()):
            if stop_flag():
                break

            row_dict = row.to_dict()
            # Convert datetime objects to strings for serialization, but preserve type info
            for key, val in row_dict.items():
                if isinstance(val, (pd.Timestamp, datetime)):
                    row_dict[key] = val.isoformat()
                elif pd.isna(val):
                    row_dict[key] = None

            task_args = (
                (i, row_dict),
                template_path,
                output_dir,
                common_column,
                file_name_column,
                other_tables,
                main_columns
            )
            tasks.append(task_args)

        if stop_flag():
            log_callback("⛔ Generation stopped by user.")
            return

        # Parallel processing with ProcessPoolExecutor
        created_docx_files = []
        failed_files = []
        start_time = time.time()

        # Use ProcessPoolExecutor for true multiprocessing
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            # Start all tasks
            future_to_index = {
                executor.submit(process_single_document, task): task[0][0]
                for task in tasks
            }

            # Process results as they complete
            completed_count = 0
            active_processes = len(future_to_index)

            log_callback(f"⚡ Actively processing {active_processes} tasks in parallel...")

            for future in as_completed(future_to_index):
                if stop_flag():
                    log_callback("⛔ Stopping all processes...")
                    # Cancel all unfinished tasks
                    for f in future_to_index:
                        f.cancel()
                    executor.shutdown(wait=False)
                    break

                try:
                    result = future.result(timeout=30)  # 30 second timeout per document
                    completed_count += 1

                    if result["success"]:
                        created_docx_files.append(result["filename"])
                        elapsed = time.time() - start_time
                        speed = completed_count / elapsed if elapsed > 0 else 0
                        log_callback(
                            f"✅ [{completed_count}/{len(tasks)}] {os.path.basename(result['filename'])} | {speed:.1f} docs/sec")
                    else:
                        error_msg = f"Row {result['index']}: {result['error']}"
                        failed_files.append(error_msg)
                        log_callback(f"❌ [{completed_count}/{len(tasks)}] {error_msg}")

                except Exception as e:
                    failed_files.append(f"Critical process error: {str(e)}")
                    log_callback(f"❌ Critical process error: {str(e)}")

        # Summary
        total_time = time.time() - start_time
        if not stop_flag():
            log_callback(f"\n🎉 Generation completed in {total_time:.1f} seconds!")
            log_callback(f"✅ Successfully created: {len(created_docx_files)} documents")
            log_callback(f"⚡ Average speed: {len(created_docx_files) / total_time:.1f} documents/second")
            if failed_files:
                log_callback(f"❌ Errors: {len(failed_files)}")
                for error in failed_files[:3]:  # Show first 3 errors
                    log_callback(f"   • {error}")
                if len(failed_files) > 3:
                    log_callback(f"   • ... and {len(failed_files) - 3} more errors")
            log_callback(f"📁 Output folder: {output_dir}")
        else:
            log_callback(
                f"\n⛔ Generation stopped after {total_time:.1f} sec. Created: {len(created_docx_files)} documents")

    except Exception as e:
        log_callback(f"❌ CRITICAL ERROR: {str(e)}")
        import traceback
        log_callback(f"Error details: {traceback.format_exc()}")